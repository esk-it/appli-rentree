"""Lecture du tableau des professeurs, où le statut est porté par la couleur.

## Ce que ce module lit

Un classeur d'une seule feuille : civilité, nom, prénom, discipline. Le
mouvement de chaque enseignant — sortant, arrivant, en congé formation,
remplacé — n'est pas écrit dans une colonne mais **peint sur sa ligne**.

C'est une pratique de tableur courante, et le fichier de l'établissement
la rend exploitable : la légende y figure, en bas, chaque couleur en
regard de son sens. Le module lit donc cette légende plutôt que de
connaître des codes hexadécimaux par cœur. Si la teinte du jaune change
l'an prochain, ou si une catégorie s'ajoute, rien à modifier ici.

## Ce qu'il refuse de faire

Une couleur absente de la légende n'est **jamais** interprétée. Elle est
signalée avec les lignes concernées, pour arbitrage humain — deviner
qu'un vert inconnu veut dire « arrivant » serait exactement le genre de
raccourci que le reste du programme s'interdit.

Une ligne sans couleur n'est pas un cas douteux : c'est un enseignant en
poste, que rien ne distingue. C'est le cas le plus fréquent.

Une ligne qui porte **deux personnes** — « CLOITRE / FUMAT », prénoms
« Morgane / Linda » — décrit un remplacement. L'ordre suggère qui est qui,
mais il ne l'affirme pas : la ligne est lue telle quelle et signalée, à
charge d'un humain de la démêler s'il faut créer deux comptes.

## Sur la fragilité de la couleur

Elle est réelle : recolorer une cellule par mégarde change un statut sans
laisser de trace, et aucune relecture ne le montrera. Le module lit donc
le fichier une fois et rend des données explicites, que l'application
conserve ensuite sous forme de texte. La couleur est une entrée, pas la
source de vérité.
"""
from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl

# Mots-clés qui rattachent un libellé de légende à un mouvement connu du
# programme. Le libellé reste conservé tel quel : c'est lui qu'on affiche.
#
# Le rapprochement se fait sur des **débuts de mots**, jamais sur des
# sous-chaînes : « part » se trouve dans « une partie de l'année », et
# classait en sortant des enseignants seulement remplacés quelques mois.
CODES = {
    "remplace": ("remplac",),
    "formation": ("formation", "stage"),
    "sortant": ("sortant", "sortie", "depart", "quitte"),
    "arrivant": ("arrivant", "arrive", "entrant", "nouveau", "nouvelle"),
}

EN_POSTE = "en_poste"
"""Aucune couleur : l'enseignant reste, rien à signaler."""


@dataclass
class EntreeLegende:
    couleur: str
    libelle: str
    code: str
    """`sortant`, `arrivant`, … ou `inconnu` si aucun mot-clé ne s'applique."""


@dataclass
class Prof:
    nom: str
    prenom: str
    discipline: str
    civilite: str
    ligne: int
    couleur: str | None
    code: str
    libelle: str


@dataclass
class RapportProfs:
    legende: list[EntreeLegende] = field(default_factory=list)
    profs: list[Prof] = field(default_factory=list)
    couleurs_hors_legende: dict[str, list[int]] = field(default_factory=dict)
    """Teinte → lignes qui la portent. Jamais interprétées."""
    notes: list[str] = field(default_factory=list)
    """Texte libre trouvé sous la légende, conservé tel quel."""
    lignes_a_deux: list[Prof] = field(default_factory=list)
    """Lignes décrivant deux personnes — un remplacement. Jamais découpées."""
    avertissements: list[str] = field(default_factory=list)

    def par_code(self, code: str) -> list[Prof]:
        return [p for p in self.profs if p.code == code]

    @property
    def nb_par_code(self) -> dict[str, int]:
        compte: dict[str, int] = {}
        for p in self.profs:
            compte[p.code] = compte.get(p.code, 0) + 1
        return compte


def _sans_accents(texte: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", texte or "")
        if unicodedata.category(c) != "Mn"
    ).lower()


def _code_depuis_libelle(libelle: str) -> str:
    """Rapproche un libellé de légende d'un mouvement connu.

    Le rapprochement se fait mot à mot, sur des débuts de mots : chercher
    « part » dans la phrase entière le trouverait aussi dans « une partie
    de l'année », et rangerait parmi les sortants un enseignant seulement
    remplacé quelques mois.
    """
    mots_du_libelle = re.split(r"[^a-z0-9]+", _sans_accents(libelle))
    for code, prefixes in CODES.items():
        if any(mot.startswith(p) for mot in mots_du_libelle for p in prefixes):
            return code
    return "inconnu"


def couleur_de(cellule) -> str | None:
    """Teinte de fond d'une cellule, sous une forme stable et comparable.

    Une couleur de thème n'a pas de valeur RVB dans le fichier — elle
    dépend du thème appliqué. On retient son index, qui suffit à
    rapprocher une ligne de sa légende puisque les deux vivent dans le
    même classeur.
    """
    remplissage = cellule.fill
    if remplissage is None or remplissage.fill_type != "solid":
        return None
    rgb = getattr(remplissage.start_color, "rgb", None)
    if isinstance(rgb, str) and rgb and rgb != "00000000":
        return "#" + rgb[-6:].upper()
    theme = getattr(remplissage.start_color, "theme", None)
    if theme is not None:
        teinte = getattr(remplissage.start_color, "tint", 0) or 0
        return f"theme{theme}" + (f"/{teinte:.2f}" if abs(teinte) > 0.005 else "")
    return None


def _couleur_ligne(cellules) -> str | None:
    """La couleur d'une ligne, prise sur la première cellule qui en porte une.

    Le tableau peint la civilité, le nom et le prénom mais laisse la
    discipline blanche : se fier à une colonne unique perdrait des lignes.
    """
    for c in cellules:
        teinte = couleur_de(c)
        if teinte:
            return teinte
    return None


def lire_fichier_profs(chemin: str | Path) -> RapportProfs:
    """Lit le classeur et rend chaque enseignant avec son mouvement.

    Raises:
        ValueError: si le classeur ne contient pas de colonne `NOM`, ou
            si la légende est introuvable — sans elle, les couleurs ne
            veulent rien dire et rien ne peut être déduit.
    """
    chemin = Path(chemin)
    if not chemin.exists():
        raise ValueError(f"Fichier introuvable : {chemin}")

    # `data_only` pour les valeurs, sans `read_only` : les styles ne sont
    # pas chargés en lecture seule, or c'est eux qui portent le sens.
    classeur = openpyxl.load_workbook(chemin, data_only=True)
    feuille = classeur[classeur.sheetnames[0]]

    rapport = RapportProfs()
    if sum(1 for _ in feuille.conditional_formatting):
        rapport.avertissements.append(
            "Le classeur contient des règles de mise en forme conditionnelle. "
            "Elles ne sont pas lues : seules les couleurs posées à la main "
            "sont interprétées."
        )

    lignes = list(feuille.iter_rows(min_row=1, max_row=feuille.max_row))

    # --- L'en-tête d'abord : est-ce seulement le bon fichier ? ------------
    depart, colonnes = None, {}
    for cellules in lignes:
        valeurs = [_sans_accents(str(c.value or "")).strip() for c in cellules]
        if "nom" in valeurs:
            depart = cellules[0].row + 1
            for i, v in enumerate(valeurs):
                if v in ("nom", "prenom", "discipline"):
                    colonnes[v] = i
            break
    if depart is None or "nom" not in colonnes:
        raise ValueError(
            "Colonne « NOM » introuvable : le classeur n'a pas la forme "
            "attendue (civilité, nom, prénom, discipline)."
        )

    # --- La légende : sans elle, aucune couleur n'a de sens ---------------
    # Elle se reconnaît à sa forme : une pastille colorée seule, suivie de
    # son libellé, et rien dans les colonnes de discipline.
    lignes_legende: set[int] = set()
    for cellules in lignes:
        if len(cellules) < 2:
            continue
        pastille = couleur_de(cellules[0])
        libelle = str(cellules[1].value or "").strip()
        reste = [str(c.value or "").strip() for c in cellules[2:]]
        if pastille and libelle and not any(reste):
            rapport.legende.append(
                EntreeLegende(
                    couleur=pastille,
                    libelle=libelle,
                    code=_code_depuis_libelle(libelle),
                )
            )
            lignes_legende.add(cellules[0].row)

    if not rapport.legende:
        raise ValueError(
            "Aucune légende trouvée dans le classeur. Sans elle, les couleurs "
            "ne peuvent pas être traduites en mouvements — et les inventer "
            "serait pire que de ne rien faire."
        )

    sens = {e.couleur: e for e in rapport.legende}
    for e in rapport.legende:
        if e.code == "inconnu":
            rapport.avertissements.append(
                f"La légende « {e.libelle} » ne correspond à aucun mouvement "
                "connu. Ses lignes sont lues, mais le programme ne sait pas "
                "quoi en faire."
            )

    def valeur(cellules, cle: str) -> str:
        i = colonnes.get(cle)
        if i is None or i >= len(cellules):
            return ""
        return str(cellules[i].value or "").strip()

    # --- Les enseignants ---------------------------------------------------
    for cellules in lignes:
        if cellules[0].row < depart or cellules[0].row in lignes_legende:
            continue
        nom = valeur(cellules, "nom")
        prenom = valeur(cellules, "prenom")
        discipline = valeur(cellules, "discipline")

        # Un enseignant porte un nom **et** au moins un prénom ou une
        # discipline. Une cellule seule dans la colonne des noms est une
        # note laissée en bas de tableau — la compter comme un enseignant
        # ajouterait une personne qui n'existe pas.
        if not nom or not (prenom or discipline):
            texte = " ".join(
                str(c.value).strip() for c in cellules if c.value is not None
            ).strip()
            if texte:
                rapport.notes.append(texte)
            continue

        teinte = _couleur_ligne(cellules)
        entree = sens.get(teinte) if teinte else None
        if teinte and entree is None:
            rapport.couleurs_hors_legende.setdefault(teinte, []).append(cellules[0].row)

        prof = Prof(
                nom=re.sub(r"\s+", " ", nom).strip(),
                prenom=re.sub(r"\s+", " ", prenom).strip(),
                discipline=discipline,
                civilite=str(cellules[0].value or "").strip(),
                ligne=cellules[0].row,
                couleur=teinte,
                code=entree.code if entree else (EN_POSTE if not teinte else "inconnu"),
                libelle=entree.libelle if entree else ("" if not teinte else "?"),
        )
        rapport.profs.append(prof)
        if "/" in prof.nom or "/" in prof.prenom:
            rapport.lignes_a_deux.append(prof)

    for teinte, lignes_concernees in rapport.couleurs_hors_legende.items():
        rapport.avertissements.append(
            f"Couleur {teinte} absente de la légende, sur {len(lignes_concernees)} "
            f"ligne(s) : {', '.join(str(l) for l in lignes_concernees[:8])}"
            + ("…" if len(lignes_concernees) > 8 else "")
            + ". Elles sont laissées sans mouvement — l'interpréter serait deviner."
        )
    if rapport.lignes_a_deux:
        exemples = ", ".join(
            f"{p.nom} ({p.ligne})" for p in rapport.lignes_a_deux[:4]
        )
        rapport.avertissements.append(
            f"{len(rapport.lignes_a_deux)} ligne(s) portent deux personnes — "
            f"un remplacement : {exemples}. Elles comptent pour une et ne sont "
            "pas découpées : l'ordre des noms suggère qui est qui, il ne "
            "l'affirme pas."
        )
    if not rapport.profs:
        rapport.avertissements.append("Aucun enseignant lu dans ce classeur.")
    return rapport
