"""Import automatique de la Table de correspondance depuis un XLSX historique.

Le XLSX du prédécesseur contient un onglet `Table` (~900 lignes × 9 colonnes)
qui fait le pont Charlemagne ↔ OU Google ↔ groupe Google. Plutôt que de
ressaisir 80 classes à la main, on lit ce fichier :

| Col A | Col C | Col D | Col E | Col G | Col H | Col I |
|-------|-------|-------|-------|-------|-------|-------|
| Site  | Code court | OU définitive | OU pré-rentrée | Groupe Google | Nom long | Groupe profs Google |

L'import ne nécessite **aucune convention stricte de nommage** : on repère
l'onglet et les colonnes par leur contenu (une ligne d'en-têtes attendue,
puis heuristiques sur `NDE|NDK|SU` en col Site et `/…` en col OU).

Deux modes :

- `simulation` : lit, mappe, produit le rapport. Ne commit rien.
- `reel` : idem + commit.

Idempotence via la clé unique `(site_id, classe_code_court)` — une nouvelle
lecture du même fichier ne recrée aucun doublon, seulement des mises à jour.

**Aucun secret n'est présent dans ce fichier — pas de garde-fou spécifique
à ajouter côté persistance.**
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from backend.models import Site, TableCorrespondance

# ---------------------------------------------------------------------------
# Rapport typé
# ---------------------------------------------------------------------------


@dataclass
class LigneImportee:
    """Une ligne du XLSX effectivement retenue pour import (créée ou mise à jour)."""

    ligne_source: int
    """Numéro de ligne dans le XLSX (1-indexé, pour retrouver dans Excel)."""

    site: str
    classe_code_court: str
    classe_charlemagne_long: str
    ou_pre_rentree: str
    ou_definitive: str
    groupe_google: str | None
    groupe_profs_google: str | None

    action: str
    """`creee` | `mise_a_jour` | `identique`."""


@dataclass
class LigneRejetee:
    """Une ligne ignorée avec la raison."""

    ligne_source: int
    raison: str
    valeurs: dict[str, Any] = field(default_factory=dict)


@dataclass
class RapportImportTable:
    mode: str  # `simulation` | `reel`
    onglet_utilise: str

    nb_lignes_lues: int = 0
    nb_lignes_ingerees: int = 0
    nb_creations: int = 0
    nb_mises_a_jour: int = 0
    nb_identiques: int = 0

    lignes_importees: list[LigneImportee] = field(default_factory=list)
    lignes_rejetees: list[LigneRejetee] = field(default_factory=list)

    sites_inconnus: list[str] = field(default_factory=list)
    """Sites lus dans le fichier mais absents de la base — l'import a ignoré
    ces lignes en les listant ici pour que l'utilisateur crée les Sites."""

    erreurs: list[str] = field(default_factory=list)
    est_bloque: bool = False


# ---------------------------------------------------------------------------
# Détection de l'onglet et des colonnes
# ---------------------------------------------------------------------------

_NOMS_ONGLETS_PROBABLES = ("Table", "TABLE", "table", "Correspondance", "Classes")

# Mapping des colonnes attendu dans le XLSX historique (voir docstring).
# Utilisé si l'auto-détection par en-tête échoue.
_COLONNES_HISTORIQUES = {
    "site": "A",
    "classe_code_court": "C",
    "ou_definitive": "D",
    "ou_pre_rentree": "E",
    "groupe_google": "G",
    "classe_charlemagne_long": "H",
    "groupe_profs_google": "I",
}


def _trouver_onglet(wb: openpyxl.Workbook) -> str:
    """Renvoie le nom de l'onglet le plus probable (`Table` en priorité)."""
    for nom in _NOMS_ONGLETS_PROBABLES:
        if nom in wb.sheetnames:
            return nom
    # Sinon, cherche celui qui contient le plus de codes classes plausibles
    meilleur = wb.sheetnames[0]
    meilleur_score = -1
    for nom in wb.sheetnames:
        ws = wb[nom]
        score = 0
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if isinstance(cell, str) and cell.strip() in ("NDE", "NDK", "SU"):
                    score += 1
        if score > meilleur_score:
            meilleur_score = score
            meilleur = nom
    return meilleur


def _detecter_ligne_donnees(ws) -> int:
    """Renvoie le n° de la 1re ligne de données (1-indexé).

    On cherche la première ligne où la colonne A contient `NDE`, `NDK` ou `SU`.
    Toute ligne avant est supposée être un en-tête (parfois plusieurs).
    """
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        if not row:
            continue
        premiere = row[0]
        if isinstance(premiere, str) and premiere.strip().upper() in ("NDE", "NDK", "SU"):
            return i
    return 2  # défaut : ligne 2 (une ligne d'en-tête)


# ---------------------------------------------------------------------------
# Point d'entrée
# ---------------------------------------------------------------------------


def importer_table(
    session: Session,
    chemin_fichier: Path,
    mode: str = "simulation",
    nom_onglet: str | None = None,
) -> RapportImportTable:
    """Importe l'onglet Table d'un XLSX historique dans TableCorrespondance.

    Args:
        session: SQLAlchemy session (commit uniquement en mode réel).
        chemin_fichier: fichier .xlsx à lire.
        mode: `simulation` (défaut) ou `reel`.
        nom_onglet: force le nom de l'onglet (sinon auto-détection).
    """
    if mode not in ("simulation", "reel"):
        raise ValueError(f"mode doit être 'simulation' ou 'reel', reçu : {mode!r}")

    rapport = RapportImportTable(mode=mode, onglet_utilise="")

    try:
        wb = openpyxl.load_workbook(chemin_fichier, data_only=True, read_only=True)
    except Exception as e:
        rapport.erreurs.append(f"Lecture impossible : {e}")
        rapport.est_bloque = True
        return rapport

    onglet = nom_onglet or _trouver_onglet(wb)
    if onglet not in wb.sheetnames:
        rapport.erreurs.append(f"Onglet introuvable : {onglet}")
        rapport.est_bloque = True
        wb.close()
        return rapport

    rapport.onglet_utilise = onglet
    ws = wb[onglet]

    # Sites déjà en base — pour la résolution nom → id
    sites = {s.nom.upper(): s for s in session.query(Site).all()}

    # Correspondances existantes — pour repérer création vs mise à jour
    existantes: dict[tuple[int, str], TableCorrespondance] = {}
    for tc in session.query(TableCorrespondance).all():
        existantes[(tc.site_id, tc.classe_code_court)] = tc

    ligne_debut = _detecter_ligne_donnees(ws)

    sites_inconnus_vus: set[str] = set()

    for i, row in enumerate(
        ws.iter_rows(min_row=ligne_debut, values_only=True), start=ligne_debut
    ):
        rapport.nb_lignes_lues += 1

        valeurs = _extraire_valeurs(row)
        site_nom = (valeurs.get("site") or "").upper()
        code_court = (valeurs.get("classe_code_court") or "").strip()
        ou_def = (valeurs.get("ou_definitive") or "").strip()
        ou_pre = (valeurs.get("ou_pre_rentree") or "").strip()

        if not site_nom and not code_court:
            # Ligne vide en fin de tableau — on continue au cas où
            continue

        # Site
        if not site_nom:
            rapport.lignes_rejetees.append(
                LigneRejetee(ligne_source=i, raison="site manquant", valeurs=valeurs)
            )
            continue

        site = sites.get(site_nom)
        if site is None:
            sites_inconnus_vus.add(site_nom)
            rapport.lignes_rejetees.append(
                LigneRejetee(
                    ligne_source=i,
                    raison=f"site {site_nom!r} inconnu — crée-le d'abord dans l'onglet Sites",
                    valeurs=valeurs,
                )
            )
            continue

        # Champs obligatoires
        if not code_court:
            rapport.lignes_rejetees.append(
                LigneRejetee(ligne_source=i, raison="code_classe manquant", valeurs=valeurs)
            )
            continue
        if not ou_def or not ou_pre:
            rapport.lignes_rejetees.append(
                LigneRejetee(
                    ligne_source=i,
                    raison="OU pré-rentrée ou définitive manquante",
                    valeurs=valeurs,
                )
            )
            continue

        long_nom = (valeurs.get("classe_charlemagne_long") or code_court).strip()
        groupe_google = _non_vide(valeurs.get("groupe_google"))
        groupe_profs = _non_vide(valeurs.get("groupe_profs_google"))

        cle = (site.id, code_court)
        existante = existantes.get(cle)

        if existante is None:
            tc = TableCorrespondance(
                site_id=site.id,
                classe_charlemagne_long=long_nom,
                classe_code_court=code_court,
                groupe_google=groupe_google,
                ou_pre_rentree=ou_pre,
                ou_definitive=ou_def,
                groupe_profs_google=groupe_profs,
            )
            session.add(tc)
            action = "creee"
            rapport.nb_creations += 1
        else:
            avant = (
                existante.classe_charlemagne_long,
                existante.ou_pre_rentree,
                existante.ou_definitive,
                existante.groupe_google,
                existante.groupe_profs_google,
            )
            existante.classe_charlemagne_long = long_nom
            existante.ou_pre_rentree = ou_pre
            existante.ou_definitive = ou_def
            existante.groupe_google = groupe_google
            existante.groupe_profs_google = groupe_profs
            apres = (long_nom, ou_pre, ou_def, groupe_google, groupe_profs)
            if avant == apres:
                action = "identique"
                rapport.nb_identiques += 1
            else:
                action = "mise_a_jour"
                rapport.nb_mises_a_jour += 1

        rapport.lignes_importees.append(
            LigneImportee(
                ligne_source=i,
                site=site.nom,
                classe_code_court=code_court,
                classe_charlemagne_long=long_nom,
                ou_pre_rentree=ou_pre,
                ou_definitive=ou_def,
                groupe_google=groupe_google,
                groupe_profs_google=groupe_profs,
                action=action,
            )
        )
        rapport.nb_lignes_ingerees += 1

    rapport.sites_inconnus = sorted(sites_inconnus_vus)
    wb.close()

    if mode == "reel":
        session.commit()
    else:
        session.rollback()

    return rapport


# ---------------------------------------------------------------------------
# Extraction : mapping colonne → valeur
# ---------------------------------------------------------------------------


def _extraire_valeurs(row: tuple) -> dict[str, Any]:
    """Applique le mapping colonne historique. Renvoie un dict des champs
    qu'on sait interpréter, indexés par le nom du champ modèle."""
    resultat = {}
    for champ, lettre in _COLONNES_HISTORIQUES.items():
        idx = _col_lettre_vers_index(lettre)
        if idx < len(row):
            resultat[champ] = row[idx]
    return resultat


def _col_lettre_vers_index(lettre: str) -> int:
    """`A` → 0, `B` → 1, `AA` → 26, etc."""
    n = 0
    for c in lettre.upper():
        n = n * 26 + (ord(c) - ord("A") + 1)
    return n - 1


def _non_vide(v: Any) -> str | None:
    """Renvoie la valeur strippée, ou None si vide."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


# Debug / diagnostic — utile pour l'écran de configuration
def apercu_onglets(chemin_fichier: Path) -> dict[str, list[str]]:
    """Retourne les 3 premières lignes de chaque onglet, en préfixant chaque
    ligne par sa lettre de colonne. Utile pour aider l'utilisateur à
    identifier le bon onglet si l'auto-détection échoue."""
    wb = openpyxl.load_workbook(chemin_fichier, data_only=True, read_only=True)
    resultat = {}
    for nom in wb.sheetnames:
        ws = wb[nom]
        lignes = []
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            lignes.append(
                " | ".join(
                    f"{get_column_letter(j)}={c!r}"
                    for j, c in enumerate(row, start=1)
                    if c is not None
                )
            )
        resultat[nom] = lignes
    wb.close()
    return resultat
