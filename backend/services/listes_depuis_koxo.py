"""Les listes et les étiquettes d'un site, tirées de son export KoXo.

## Pourquoi cet export-là et pas le référentiel

Le référentiel connaît le nom, la classe, le login et l'adresse. Il ne
connaît **pas le mot de passe** — et c'est délibéré : là où KoXo existe,
c'est lui qui en est l'autorité, le programme n'en invente aucun.

Or les trois documents qu'on distribue à la rentrée en ont tous besoin :
la liste que le professeur principal garde sous la main, celle des
entrants qu'on remet à la vie scolaire, et les étiquettes que l'élève
emporte. Les tirer du référentiel donnerait des colonnes vides.

Ils se tirent donc de l'**export KoXo avec les mots de passe** — le même
fichier qui sert déjà à remplir le mot de passe des comptes Google.

## Ce que le programme apporte

KoXo sait déjà imprimer ses fiches. Ce qu'il ne sait pas faire :

- **distinguer les entrants** — il ignore l'année précédente, quand le
  référentiel la photographie ;
- **rendre un classeur** que l'on trie et que l'on filtre.

Le croisement se fait sur l'`ID unique`, où le programme écrit toujours le
badge Charlemagne. Jamais sur le nom : deux homonymes suffiraient à
échanger deux mots de passe.

## Les mots de passe ne sont pas conservés

Ils traversent la mémoire, entrent dans les fichiers rendus, et rien
n'est écrit en base. Le coffre sert aux mots de passe **que le programme
fabrique** — ceux de NDE, qui n'a pas de KoXo et où les perdre obligerait
à réinitialiser chaque compte. Ici l'autorité est KoXo : il les redonnera.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field

from sqlalchemy.orm import Session

from backend.models import AnneeScolaire, Personne, Site, Snapshot
from backend.services.rattachement import ids_personnes_du_site, ids_presents_annee
from backend.services.regles_metier import classe_lisible

COLONNES = ("Nom", "Prénom", "Classe", "Identifiant", "Mot de passe", "Adresse")


class ListesImpossibles(Exception):
    """La génération est refusée, et le message dit pourquoi."""


@dataclass
class LigneListe:
    personne_id: int
    nom: str
    prenom: str
    classe: str
    login: str
    mot_de_passe: str
    email: str
    nouveau: bool = False

    @property
    def classe_affichee(self) -> str:
        """La classe telle qu'on l'imprime, pas telle qu'on la stocke."""
        return classe_lisible(self.classe)


@dataclass
class RapportListes:
    site_nom: str
    annee_libelle: str

    lignes: list[LigneListe] = field(default_factory=list)
    nouveaux: list[LigneListe] = field(default_factory=list)

    sans_ligne_koxo: list[str] = field(default_factory=list)
    """Inscrits au site que l'export KoXo ne porte pas — donc sans mot de
    passe à distribuer."""
    sans_mot_de_passe: list[str] = field(default_factory=list)
    """Présents dans l'export, mais la colonne du mot de passe est vide :
    l'export a été pris sans cocher « inclure les mots de passe »."""
    koxo_hors_site: int = 0
    """Lignes de l'export qui ne concernent pas ce site — un autre serveur,
    ou des comptes que le référentiel ne connaît pas."""

    xlsx_tous: bytes = b""
    xlsx_nouveaux: bytes = b""
    etiquettes_nouveaux: bytes = b""
    nom_xlsx_tous: str = ""
    nom_xlsx_nouveaux: str = ""
    nom_etiquettes: str = ""

    @property
    def nb_tous(self) -> int:
        return len(self.lignes)

    @property
    def nb_nouveaux(self) -> int:
        return len(self.nouveaux)


def listes_depuis_koxo(
    session: Session,
    lignes_koxo: list,
    *,
    site_id: int,
    annee_cible_id: int,
    annee_source_id: int | None = None,
) -> RapportListes:
    """Trois documents d'un seul export : la liste, les entrants, les fiches.

    Args:
        lignes_koxo: les lignes de `lire_export_brut` — pas le tuple entier.
        annee_source_id: sans elle, on ne peut pas dire qui entre. La liste
            des nouveaux et leurs étiquettes ne sont alors pas produites,
            plutôt que rendues fausses.

    Raises:
        ListesImpossibles: site ou année inconnus, export vide, ou aucun
            élève du site retrouvé dedans.
    """
    site = session.query(Site).filter_by(id=site_id).one_or_none()
    if site is None:
        raise ListesImpossibles(f"Site introuvable : {site_id}")
    annee = session.query(AnneeScolaire).filter_by(id=annee_cible_id).one_or_none()
    if annee is None:
        raise ListesImpossibles(f"Année introuvable : {annee_cible_id}")
    if not lignes_koxo:
        raise ListesImpossibles("L'export KoXo ne contient aucune ligne.")
    if annee_source_id is not None and annee_source_id == annee_cible_id:
        # Comparer une année à elle-même ne rend aucun entrant, et le
        # rendait sans rien dire : trois documents produits, celui des
        # nouveaux vide, et aucune raison visible. L'écran classait les
        # années par date de création, où « 2025-2026 » venait après
        # « 2026-2027 » — la source valait alors la cible.
        raise ListesImpossibles(
            f"L'année source et l'année cible sont la même ({annee.libelle}) : "
            "aucun élève ne peut y être « nouveau ». Choisis l'année "
            "précédente comme source."
        )

    par_badge = {}
    for l in lignes_koxo:
        ident = (getattr(l, "id_unique", "") or "").strip()
        if ident:
            par_badge[ident] = l

    ids_site = ids_personnes_du_site(
        session, site_id=site_id, annee_id=annee_cible_id, type_personne="eleve"
    )
    classes = _classes_de_lannee(session, annee_cible_id)
    anciens = (
        ids_presents_annee(session, annee_id=annee_source_id, type_personne="eleve")
        if annee_source_id is not None
        else set()
    )

    rapport = RapportListes(site_nom=site.nom, annee_libelle=annee.libelle)
    vus_dans_koxo: set[str] = set()

    personnes = (
        session.query(Personne)
        .filter(Personne.id.in_(ids_site), Personne.type == "eleve")
        .all()
    )
    for p in sorted(personnes, key=lambda x: ((classes.get(x.id) or ""), x.nom, x.prenom)):
        badge = str(p.badge) if p.badge is not None else ""
        k = par_badge.get(badge)
        qui = f"{p.prenom} {p.nom}"
        if k is None:
            rapport.sans_ligne_koxo.append(qui)
            continue
        vus_dans_koxo.add(badge)
        mdp = (getattr(k, "mot_de_passe", "") or "").strip()
        if not mdp:
            rapport.sans_mot_de_passe.append(qui)

        ligne = LigneListe(
            personne_id=p.id,
            nom=p.nom,
            prenom=p.prenom,
            classe=classes.get(p.id) or (p.classe or ""),
            # Le login de KoXo fait foi : c'est celui avec lequel l'élève se
            # connecte au réseau, et il peut différer de celui du
            # référentiel après une reprise manuelle.
            login=(getattr(k, "login", "") or "").strip() or (p.login or ""),
            mot_de_passe=mdp,
            email=(p.email or ""),
            nouveau=annee_source_id is not None and p.id not in anciens,
        )
        rapport.lignes.append(ligne)
        if ligne.nouveau:
            rapport.nouveaux.append(ligne)

    rapport.koxo_hors_site = len(par_badge) - len(vus_dans_koxo)
    if not rapport.lignes:
        raise ListesImpossibles(
            f"Aucun élève de {site.nom} n'a été retrouvé dans cet export KoXo. "
            "C'est probablement l'export d'un autre serveur — KoXo a une base "
            "par établissement."
        )

    _composer(rapport, site, annee, avec_nouveaux=annee_source_id is not None)
    return rapport


def _composer(rapport: RapportListes, site, annee, *, avec_nouveaux: bool) -> None:
    from backend.services.comptes_sans_koxo import fiches_html

    rapport.xlsx_tous = _classeur(rapport.lignes, f"{site.nom} {annee.libelle}")
    rapport.nom_xlsx_tous = f"Comptes_{site.nom}_{annee.libelle}_tous.xlsx"

    if not avec_nouveaux:
        # Sans année précédente, « nouveau » n'a pas de sens : ne rien rendre
        # vaut mieux qu'un fichier où tout le monde serait entrant.
        return

    rapport.xlsx_nouveaux = _classeur(rapport.nouveaux, f"{site.nom} entrants")
    rapport.nom_xlsx_nouveaux = f"Comptes_{site.nom}_{annee.libelle}_nouveaux.xlsx"

    rapport.etiquettes_nouveaux = _etiquettes(rapport, site, annee)
    rapport.nom_etiquettes = f"Etiquettes_{site.nom}_{annee.libelle}_nouveaux.html"


def _etiquettes(rapport: RapportListes, site, annee) -> bytes:
    from backend.services.comptes_sans_koxo import fiches_html

    return fiches_html(
        [
            {
                "nom": l.nom, "prenom": l.prenom,
                "classe": l.classe_affichee,
                "groupe": l.classe_affichee, "login": l.login,
                "mot_de_passe": l.mot_de_passe,
                # La ligne « Email » de l'étiquette lit `adresse`. L'oublier
                # ne lève rien : l'élève repartait avec « Email : » suivi
                # de rien, et c'est justement l'adresse qu'il vient chercher.
                "adresse": l.email,
            }
            for l in rapport.nouveaux
        ],
        # Le bandeau nomme l'établissement — « Collège Sainte Ursule » — et
        # non l'OGEC qui le gère : c'est un papier remis à l'élève, qui
        # reconnaît son collège, pas son organisme gestionnaire.
        organisation=site.nom_complet or site.organisation_etiquettes or site.nom,
        annee=annee.libelle,
        # Là où KoXo existe, l'élève ouvre aussi une session sur le réseau :
        # les mêmes identifiants servent deux fois, et l'étiquette le dit.
        avec_reseau=bool(site.base_koxo),
    )


def _classeur(lignes: list[LigneListe], titre_feuille: str) -> bytes:
    """Un classeur qu'on trie et qu'on filtre — pas un tableau figé.

    L'en-tête est figé et le filtre posé : ces listes se lisent classe par
    classe, et devoir reposer le filtre à chaque ouverture est le genre de
    détail qui fait qu'on ne s'en sert pas.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titre_feuille[:31]

    ws.append(list(COLONNES))
    gras = Font(bold=True, color="FFFFFF")
    fond = PatternFill("solid", fgColor="047857")
    for i in range(1, len(COLONNES) + 1):
        c = ws.cell(row=1, column=i)
        c.font, c.fill = gras, fond
        c.alignment = Alignment(vertical="center")

    for l in lignes:
        ws.append([l.nom, l.prenom, l.classe_affichee, l.login,
                   l.mot_de_passe, l.email])

    largeurs = (24, 18, 12, 14, 14, 34)
    for i, w in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    if lignes:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLONNES))}{len(lignes) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _classes_de_lannee(session: Session, annee_id: int) -> dict[int, str | None]:
    """La classe de l'année photographiée, pas la classe courante.

    Les deux divergent dès qu'un mouvement a lieu après l'ingestion, et
    c'est la photographie qui fait foi pour un document daté.
    """
    derniers: dict[int, str | None] = {}
    for s in (
        session.query(Snapshot)
        .filter(Snapshot.annee_scolaire_id == annee_id)
        .order_by(Snapshot.personne_id, Snapshot.date_ingestion.desc())
        .all()
    ):
        derniers.setdefault(s.personne_id, s.classe)
    return derniers
