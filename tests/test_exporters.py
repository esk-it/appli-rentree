"""Tests des exporters (vérifie le format des fichiers générés)."""
from __future__ import annotations

import base64
import csv
import io

import pytest
from openpyxl import load_workbook

from backend.services.exporters.cardstudio import generer_exports_cardstudio
from backend.services.exporters.google import generer_exports_google
from backend.services.exporters.koxo import generer_exports_koxo
from backend.services.exporters.pmb import generer_exports_pmb
from backend.services.exporters.smartair import (
    generer_exports_smartair,
    parser_export_smartair_n_minus_1,
)


@pytest.fixture()
def snapshot_simple(session, etablissement_factory, annee_factory, eleve_factory):
    """Snapshot avec 5 élèves SU et 3 NDK_LY."""
    etab_su = etablissement_factory("02-COL", "SU", "Collège SU", "college")
    etab_ndk = etablissement_factory(
        "03-LY", "NDK_LY", "L.E.G.T. NDK", "lycee_general"
    )
    annee = annee_factory("2025-2026")
    for i in range(5):
        eleve_factory(
            annee_id=annee.id,
            etablissement_id=etab_su.id,
            num_badge=50000 + i,
            nom=f"SU_NOM{i:02d}",
            code_classe="31",
        )
    for i in range(3):
        eleve_factory(
            annee_id=annee.id,
            etablissement_id=etab_ndk.id,
            num_badge=60000 + i,
            nom=f"NDK_NOM{i:02d}",
            code_classe="2_5",
        )
    return {"annee": annee, "etab_su": etab_su, "etab_ndk": etab_ndk}


class TestKoXo:
    def test_genere_un_fichier_par_groupe(self, snapshot_simple, session):
        fichiers = generer_exports_koxo(session, "2025-2026", seed=42)
        noms = {f.nom for f in fichiers}
        assert "KoXo_SU_Eleves_Tous_2025-2026.csv" in noms
        assert "KoXo_NDK_Eleves_Tous_2025-2026.csv" in noms

    def test_nb_lignes_correspond(self, snapshot_simple, session):
        fichiers = generer_exports_koxo(session, "2025-2026")
        su = next(f for f in fichiers if "SU" in f.nom)
        ndk = next(f for f in fichiers if "NDK" in f.nom)
        assert su.nb_lignes == 5
        assert ndk.nb_lignes == 3

    def test_csv_a_les_bonnes_colonnes(self, snapshot_simple, session):
        fichiers = generer_exports_koxo(session, "2025-2026")
        contenu = fichiers[0].contenu.lstrip("﻿")
        reader = csv.DictReader(io.StringIO(contenu))
        ligne = next(reader)
        assert set(ligne.keys()) == {
            "Groupe primaire",
            "Groupe secondaire",
            "Titre",
            "Nom",
            "Prénom",
            "Identifiant",
            "ID unique",
            "Mot de passe",
            "Date de naissance",
            "Email",
        }
        assert ligne["Groupe primaire"] == "Elèves"

    def test_avec_n_minus_1_genere_nouveaux_et_anciens(
        self, session, etablissement_factory, annee_factory, eleve_factory
    ):
        etab = etablissement_factory()
        n1 = annee_factory("2024-2025")
        n = annee_factory("2025-2026")
        eleve_factory(annee_id=n1.id, etablissement_id=etab.id, num_badge=100)  # sortant
        eleve_factory(annee_id=n.id, etablissement_id=etab.id, num_badge=100)  # restant
        # Wait: on a un seul badge dans les deux → restant. Pour avoir un sortant il
        # faut un badge présent en N-1 mais pas en N.
        eleve_factory(annee_id=n1.id, etablissement_id=etab.id, num_badge=200)
        # En N, on ajoute un nouveau
        eleve_factory(annee_id=n.id, etablissement_id=etab.id, num_badge=300)

        fichiers = generer_exports_koxo(session, "2025-2026", "2024-2025", seed=1)
        noms = {f.nom for f in fichiers}
        assert any("Nouveaux" in n for n in noms)
        assert any("Anciens" in n for n in noms)


class TestPmb:
    def test_un_fichier_par_instance(self, snapshot_simple, session):
        fichiers = generer_exports_pmb(session, "2025-2026")
        noms = {f.nom for f in fichiers}
        assert any("SU" in n for n in noms)
        assert any("NDK" in n for n in noms)

    def test_separateur_pointvirgule(self, snapshot_simple, session):
        fichiers = generer_exports_pmb(session, "2025-2026")
        contenu = fichiers[0].contenu.lstrip("﻿")
        # PMB doit utiliser ; pas ,
        premiere_ligne = contenu.split("\r\n")[0]
        assert ";" in premiere_ligne
        assert premiere_ligne.count(";") >= 5


class TestCardStudio:
    def test_xlsx_lisible(self, snapshot_simple, session):
        fichiers = generer_exports_cardstudio(session, "2025-2026")
        assert len(fichiers) >= 1
        xlsx_bytes = base64.b64decode(fichiers[0].contenu_base64)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        # En-tête
        headers = [c.value for c in ws[1]]
        assert "Etablissement" in headers
        assert "Num Badge" in headers
        assert "NomFichierPhoto" in headers

    def test_nb_lignes_dans_xlsx(self, snapshot_simple, session):
        fichiers = generer_exports_cardstudio(session, "2025-2026")
        su_fichier = next(f for f in fichiers if "SAINTE-URSULE" in f.nom)
        xlsx_bytes = base64.b64decode(su_fichier.contenu_base64)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        # 5 élèves + 1 ligne d'en-tête = 6
        assert ws.max_row == 6


class TestSmartAir:
    def test_genere_avec_op_a_par_defaut(self, snapshot_simple, session):
        fichiers = generer_exports_smartair(session, "2025-2026")
        assert len(fichiers) == 1
        contenu = fichiers[0].contenu.lstrip("﻿")
        reader = csv.DictReader(io.StringIO(contenu), delimiter=";")
        ops = [l["Op"] for l in reader]
        assert all(op == "a" for op in ops)

    def test_avec_n_minus_1_distingue_a_m_b(
        self, session, etablissement_factory, annee_factory, eleve_factory
    ):
        etab = etablissement_factory()
        # Préparer le N
        n = annee_factory("2025-2026")
        eleve_factory(annee_id=n.id, etablissement_id=etab.id, num_badge=111)  # m
        eleve_factory(annee_id=n.id, etablissement_id=etab.id, num_badge=222)  # a

        badges_n_1 = {111, 333}  # 333 absent de N → b
        card_ids = {111: "CARDABC", 333: "CARDXYZ"}
        fichiers = generer_exports_smartair(
            session, "2025-2026", card_ids_existants=card_ids, badges_n_minus_1=badges_n_1
        )
        contenu = fichiers[0].contenu.lstrip("﻿")
        lignes = list(csv.DictReader(io.StringIO(contenu), delimiter=";"))
        ops_par_id = {l["Id"]: l["Op"] for l in lignes}
        assert ops_par_id["111"] == "m"
        assert ops_par_id["222"] == "a"
        assert ops_par_id["333"] == "b"
        # CardId préservé pour 111 (présent en N-1) et 333 (sortant)
        cards = {l["Id"]: l["CardId"] for l in lignes}
        assert cards["111"] == "CARDABC"
        assert cards["333"] == "CARDXYZ"
        assert cards["222"] == ""  # nouveau, pas de CardId

    def test_parser_export_smartair_n_minus_1(self):
        contenu = (
            "Op;Id;Name;CardId;Group\r\n"
            ";100;DOE John;ABC123;3A\r\n"
            ";200;DOE Jane;XYZ789;3B\r\n"
        )
        card_ids, badges = parser_export_smartair_n_minus_1(contenu)
        assert card_ids == {100: "ABC123", 200: "XYZ789"}
        assert badges == {100, 200}


class TestGoogle:
    def test_ou_path_par_defaut(self, snapshot_simple, session):
        fichiers = generer_exports_google(session, "2025-2026")
        contenu = fichiers[0].contenu.lstrip("﻿")
        reader = csv.DictReader(io.StringIO(contenu))
        ligne = next(reader)
        ou = ligne["Org Unit Path [Required]"]
        # Pattern par défaut : /{site}/{site}{annee}/{classe}
        assert ou.startswith("/")
        # Doit contenir SU et 2026
        assert "SU" in ou or "NDK" in ou
        assert "2026" in ou

    def test_email_calcule_dans_csv(self, snapshot_simple, session):
        fichiers = generer_exports_google(session, "2025-2026")
        contenu = fichiers[0].contenu.lstrip("﻿")
        reader = csv.DictReader(io.StringIO(contenu))
        ligne = next(reader)
        assert "@lekreisker.fr" in ligne["Email Address [Required]"]
