"""Tests d'import de la Table de correspondance depuis un XLSX historique.

Le fichier XLSX du prédécesseur mappe :
| A | (B) | C | D | E | (F) | G | H | I |
| Site | base_path | code_court | OU_def | OU_pre | site_bis | groupe | classe_long | groupe_profs |

Ces tests construisent des XLSX à la volée avec openpyxl.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest


# ---------------------------------------------------------------------------
# Fabrique de XLSX en RAM
# ---------------------------------------------------------------------------


def _ecrire_xlsx(chemin: Path, lignes: list[list], nom_onglet: str = "Table") -> None:
    """Crée un XLSX avec l'onglet nommé et les lignes fournies (1re = en-tête)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nom_onglet
    for row in lignes:
        ws.append(row)
    wb.save(chemin)


def _ligne_type(site, code_court, ou_def, ou_pre, groupe, classe_long, groupe_profs):
    """Construit une ligne au format historique (9 colonnes)."""
    return [site, "", code_court, ou_def, ou_pre, site, groupe, classe_long, groupe_profs]


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_import_creation_basique(tmp_path, session, site_factory):
    """Un XLSX minimal crée les lignes attendues en mode réel."""
    from backend.models import TableCorrespondance
    from backend.services.import_table import importer_table

    ndk = site_factory("NDK")
    fic = tmp_path / "table.xlsx"
    _ecrire_xlsx(fic, [
        ["Site", "Base", "Code", "OU def", "OU pre", "Site", "Groupe", "Long", "Profs"],
        _ligne_type("NDK", "3B", "/3. NDK/NDK2026/3B", "/3. NDK/NDK2026",
                    "3eme-b@lekreisker.fr", "TROISIEME B", "profs-3b@lekreisker.fr"),
        _ligne_type("NDK", "4A", "/3. NDK/NDK2026/4A", "/3. NDK/NDK2026",
                    "4eme-a@lekreisker.fr", "QUATRIEME A", "profs-4a@lekreisker.fr"),
    ])

    r = importer_table(session, fic, mode="reel")

    assert r.est_bloque is False
    assert r.nb_lignes_lues == 2
    assert r.nb_creations == 2
    assert r.nb_lignes_ingerees == 2

    tcs = session.query(TableCorrespondance).order_by(TableCorrespondance.classe_code_court).all()
    assert [t.classe_code_court for t in tcs] == ["3B", "4A"]
    assert tcs[0].ou_definitive == "/3. NDK/NDK2026/3B"
    assert tcs[0].groupe_google == "3eme-b@lekreisker.fr"
    assert tcs[0].site_id == ndk.id


def test_import_simulation_ne_persiste_rien(tmp_path, session, site_factory):
    from backend.models import TableCorrespondance
    from backend.services.import_table import importer_table

    site_factory("NDK")
    fic = tmp_path / "sim.xlsx"
    _ecrire_xlsx(fic, [
        ["Site", "Base", "Code", "OU def", "OU pre", "Site", "Groupe", "Long", "Profs"],
        _ligne_type("NDK", "3B", "/3. NDK/NDK2026/3B", "/3. NDK/NDK2026",
                    "3eme-b@lekreisker.fr", "TROISIEME B", ""),
    ])

    r = importer_table(session, fic, mode="simulation")

    assert r.nb_creations == 1  # dans le rapport
    assert session.query(TableCorrespondance).count() == 0  # mais rien en base


def test_import_site_inconnu_est_rejete(tmp_path, session, site_factory):
    from backend.services.import_table import importer_table

    # On ne crée que NDK — NDE devra apparaître dans sites_inconnus
    site_factory("NDK")

    fic = tmp_path / "site_inconnu.xlsx"
    _ecrire_xlsx(fic, [
        ["Site", "Base", "Code", "OU def", "OU pre", "Site", "Groupe", "Long", "Profs"],
        _ligne_type("NDK", "3B", "/3. NDK/NDK2026/3B", "/3. NDK/NDK2026",
                    "", "TROISIEME B", ""),
        _ligne_type("NDE", "3F", "/2. NDE/NDE2026/3F", "/2. NDE/NDE2026",
                    "3eme-fuschia@ndecleder.fr", "TROISIEME FUSHIA", ""),
    ])

    r = importer_table(session, fic, mode="reel")

    assert r.nb_creations == 1  # seul NDK/3B
    assert "NDE" in r.sites_inconnus
    assert any("NDE" in lr.raison and "inconnu" in lr.raison for lr in r.lignes_rejetees)


def test_import_idempotent_relance(tmp_path, session, site_factory):
    """Deux imports successifs du même fichier : la 2e ne fait que des identiques."""
    from backend.models import TableCorrespondance
    from backend.services.import_table import importer_table

    site_factory("NDK")
    fic = tmp_path / "table.xlsx"
    _ecrire_xlsx(fic, [
        ["Site", "Base", "Code", "OU def", "OU pre", "Site", "Groupe", "Long", "Profs"],
        _ligne_type("NDK", "3B", "/3. NDK/NDK2026/3B", "/3. NDK/NDK2026",
                    "3eme-b@lekreisker.fr", "TROISIEME B", ""),
    ])

    r1 = importer_table(session, fic, mode="reel")
    assert r1.nb_creations == 1

    r2 = importer_table(session, fic, mode="reel")
    assert r2.nb_creations == 0
    assert r2.nb_identiques == 1
    assert r2.nb_mises_a_jour == 0

    # Un seul enregistrement en base
    assert session.query(TableCorrespondance).count() == 1


def test_import_detecte_mise_a_jour(tmp_path, session, site_factory):
    """Si l'OU définitive change dans le XLSX, la ligne existante est mise à jour."""
    from backend.models import TableCorrespondance
    from backend.services.import_table import importer_table

    site_factory("NDK")
    fic1 = tmp_path / "v1.xlsx"
    _ecrire_xlsx(fic1, [
        ["Site", "Base", "Code", "OU def", "OU pre", "Site", "Groupe", "Long", "Profs"],
        _ligne_type("NDK", "3B", "/3. NDK/NDK2025/3B", "/3. NDK/NDK2025",
                    "3eme-b@lekreisker.fr", "TROISIEME B", ""),
    ])
    importer_table(session, fic1, mode="reel")

    # Nouveau XLSX avec l'OU 2026 au lieu de 2025
    fic2 = tmp_path / "v2.xlsx"
    _ecrire_xlsx(fic2, [
        ["Site", "Base", "Code", "OU def", "OU pre", "Site", "Groupe", "Long", "Profs"],
        _ligne_type("NDK", "3B", "/3. NDK/NDK2026/3B", "/3. NDK/NDK2026",
                    "3eme-b@lekreisker.fr", "TROISIEME B", ""),
    ])
    r = importer_table(session, fic2, mode="reel")

    assert r.nb_mises_a_jour == 1
    assert r.nb_creations == 0

    tc = session.query(TableCorrespondance).filter_by(classe_code_court="3B").one()
    assert tc.ou_definitive == "/3. NDK/NDK2026/3B"


def test_import_ignore_ligne_sans_code(tmp_path, session, site_factory):
    from backend.services.import_table import importer_table

    site_factory("NDK")
    fic = tmp_path / "manquant.xlsx"
    _ecrire_xlsx(fic, [
        ["Site", "Base", "Code", "OU def", "OU pre", "Site", "Groupe", "Long", "Profs"],
        _ligne_type("NDK", "", "/3. NDK/NDK2026/3B", "/3. NDK/NDK2026",
                    "3eme-b@lekreisker.fr", "TROISIEME B", ""),
    ])

    r = importer_table(session, fic, mode="reel")

    assert r.nb_creations == 0
    assert any("code_classe" in lr.raison for lr in r.lignes_rejetees)


def test_import_ignore_ligne_sans_ou(tmp_path, session, site_factory):
    from backend.services.import_table import importer_table

    site_factory("NDK")
    fic = tmp_path / "sans_ou.xlsx"
    _ecrire_xlsx(fic, [
        ["Site", "Base", "Code", "OU def", "OU pre", "Site", "Groupe", "Long", "Profs"],
        _ligne_type("NDK", "3B", "", "", "3eme-b@lekreisker.fr", "TROISIEME B", ""),
    ])

    r = importer_table(session, fic, mode="reel")
    assert r.nb_creations == 0
    assert any("OU" in lr.raison for lr in r.lignes_rejetees)


def test_mode_invalide_leve_valueerror(tmp_path, session):
    from backend.services.import_table import importer_table

    fic = tmp_path / "vide.xlsx"
    _ecrire_xlsx(fic, [["Site"]])
    with pytest.raises(ValueError, match="mode"):
        importer_table(session, fic, mode="fantaisie")


def test_lecture_fichier_invalide(tmp_path, session):
    from backend.services.import_table import importer_table

    fic = tmp_path / "pas_un_xlsx.xlsx"
    fic.write_text("ceci n'est pas un xlsx")
    r = importer_table(session, fic, mode="simulation")

    assert r.est_bloque is True
    assert len(r.erreurs) == 1


def test_apercu_onglets(tmp_path):
    from backend.services.import_table import apercu_onglets

    fic = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Table"
    wb["Table"].append(["Site", "Base", "Code"])
    wb["Table"].append(["NDK", "", "3B"])
    wb.create_sheet("Autre").append(["truc"])
    wb.save(fic)

    resultat = apercu_onglets(fic)
    assert "Table" in resultat
    assert "Autre" in resultat
    assert any("NDK" in ligne for ligne in resultat["Table"])


def test_gros_import_avec_plusieurs_sites(tmp_path, session, site_factory):
    """Simule un import d'onglet réaliste : 3 sites × plusieurs classes chacun."""
    from backend.models import TableCorrespondance
    from backend.services.import_table import importer_table

    site_factory("NDE")
    site_factory("NDK")
    site_factory("SU")

    lignes = [["Site", "Base", "Code", "OU def", "OU pre", "Site", "Groupe", "Long", "Profs"]]
    for i in range(1, 4):
        lignes.append(_ligne_type("NDE", f"3F{i}", f"/2. NDE/NDE2026/3F{i}",
                                   "/2. NDE/NDE2026", f"3eme-{i}@ndecleder.fr", f"TROISIEME {i}", ""))
    for i in range(1, 6):
        lignes.append(_ligne_type("NDK", f"2ND{i}", f"/3. NDK/NDK2026/2ND{i}",
                                   "/3. NDK/NDK2026", f"2nde-{i}@lekreisker.fr", f"SECONDE {i}", ""))
    for i in range(1, 4):
        lignes.append(_ligne_type("SU", f"6{i}", f"/4. SU/SU2026/6{i}",
                                   "/4. SU/SU2026", f"6eme-{i}@lekreisker.fr", f"SIXIEME {i}", ""))

    fic = tmp_path / "reel.xlsx"
    _ecrire_xlsx(fic, lignes)

    r = importer_table(session, fic, mode="reel")

    assert r.nb_creations == 11
    assert r.nb_lignes_ingerees == 11
    assert session.query(TableCorrespondance).count() == 11
