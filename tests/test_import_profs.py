"""Lecture du tableau des professeurs, dont le statut est porté par la couleur."""
from __future__ import annotations

import openpyxl
import pytest
from openpyxl.styles import PatternFill

JAUNE = "FFFFFF00"
BLEU = "FF00B0F0"
VERT = "FF00B050"


def _classeur(tmp_path, lignes, legende, nom="profs.xlsx"):
    """Fabrique un classeur de la même forme que celui de l'établissement.

    `lignes` : (civilite, nom, prenom, discipline, couleur | None)
    `legende` : (couleur, libellé), posée sous le tableau.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None, "NOM", "PRENOM", "DISCIPLINE"])
    for civ, n, p, d, couleur in lignes:
        ws.append([civ, n, p, d])
        if couleur:
            remplissage = PatternFill(start_color=couleur, end_color=couleur,
                                      fill_type="solid")
            for col in "ABC":
                ws[f"{col}{ws.max_row}"].fill = remplissage
    ws.append([])
    for couleur, libelle in legende:
        ws.append([None, libelle])
        ws[f"A{ws.max_row}"].fill = PatternFill(
            start_color=couleur, end_color=couleur, fill_type="solid"
        )
    chemin = tmp_path / nom
    wb.save(chemin)
    return chemin


LEGENDE = [(JAUNE, "Profs sortants"), (BLEU, "Profs arrivants")]


def test_la_couleur_donne_le_mouvement(tmp_path):
    from backend.services.import_profs import lire_fichier_profs

    f = _classeur(tmp_path, [
        ("M.", "ALDRIN", "Thierry", "Technologie", JAUNE),
        ("Mme", "CAREME", "Mégane", "Maths", BLEU),
        ("M.", "SERIO", "Philippe", "Technologie", None),
    ], LEGENDE)

    r = lire_fichier_profs(f)
    assert r.nb_par_code == {"sortant": 1, "arrivant": 1, "en_poste": 1}
    assert r.par_code("sortant")[0].nom == "ALDRIN"
    assert r.par_code("sortant")[0].libelle == "Profs sortants"
    assert r.avertissements == []


def test_la_legende_vient_du_fichier_pas_du_code(tmp_path):
    """Changer la teinte ou renommer une catégorie ne doit rien casser."""
    from backend.services.import_profs import lire_fichier_profs

    f = _classeur(tmp_path, [
        ("M.", "X", "Un", "Maths", VERT),
    ], [(VERT, "Départs de fin d'année")])

    r = lire_fichier_profs(f)
    assert r.legende[0].couleur == "#00B050"
    assert r.par_code("sortant")[0].nom == "X"


def test_une_partie_de_lannee_nest_pas_un_depart(tmp_path):
    """« part » se trouve dans « une partie de l'année ».

    Le rapprochement par sous-chaîne rangeait parmi les sortants un
    enseignant seulement remplacé quelques mois — et son Chromebook aurait
    été réclamé à tort.
    """
    from backend.services.import_profs import lire_fichier_profs

    f = _classeur(tmp_path, [
        ("M.", "X", "Un", "Maths", VERT),
    ], [(VERT, "Profs remplacé sur une partie de l'année")])

    r = lire_fichier_profs(f)
    assert r.legende[0].code == "remplace"
    assert r.par_code("sortant") == []


def test_une_couleur_hors_legende_nest_jamais_interpretee(tmp_path):
    from backend.services.import_profs import lire_fichier_profs

    f = _classeur(tmp_path, [
        ("M.", "CONNU", "Un", "Maths", JAUNE),
        ("Mme", "MYSTERE", "Deux", "SVT", VERT),
    ], LEGENDE)

    r = lire_fichier_profs(f)
    mystere = [p for p in r.profs if p.nom == "MYSTERE"][0]
    assert mystere.code == "inconnu"
    assert "#00B050" in r.couleurs_hors_legende
    assert any("absente de la légende" in a for a in r.avertissements)


def test_sans_legende_le_fichier_est_refuse(tmp_path):
    """Sans elle, les couleurs ne veulent rien dire — les inventer serait pire."""
    from backend.services.import_profs import lire_fichier_profs

    f = _classeur(tmp_path, [("M.", "X", "Un", "Maths", JAUNE)], [])
    with pytest.raises(ValueError, match="légende"):
        lire_fichier_profs(f)


def test_une_note_en_bas_nest_pas_un_enseignant(tmp_path):
    """Le fichier réel finit par une note libre dans la colonne des noms."""
    from backend.services.import_profs import lire_fichier_profs

    f = _classeur(tmp_path, [("M.", "X", "Un", "Maths", None)], LEGENDE)
    wb = openpyxl.load_workbook(f)
    ws = wb.active
    ws.append([])
    ws.append([None, 'prévoir un chromebook "PROF1"…'])
    wb.save(f)

    r = lire_fichier_profs(f)
    assert [p.nom for p in r.profs] == ["X"]
    assert any("PROF1" in n for n in r.notes)


def test_une_ligne_a_deux_personnes_est_signalee(tmp_path):
    """Un remplacement tient sur une ligne ; l'ordre suggère, il n'affirme pas."""
    from backend.services.import_profs import lire_fichier_profs

    f = _classeur(tmp_path, [
        ("Mme", "CLOITRE / FUMAT", "Morgane / Linda", "Breton", VERT),
    ], [(VERT, "Profs remplacé sur une partie de l'année")])

    r = lire_fichier_profs(f)
    assert len(r.profs) == 1, "la ligne compte pour une, sans découpage"
    assert len(r.lignes_a_deux) == 1
    assert any("deux personnes" in a for a in r.avertissements)


def test_la_couleur_est_lue_meme_si_seule_la_civilite_est_peinte(tmp_path):
    """Le tableau laisse parfois la discipline blanche."""
    from backend.services.import_profs import lire_fichier_profs

    f = _classeur(tmp_path, [("M.", "X", "Un", "Maths", None)], LEGENDE)
    wb = openpyxl.load_workbook(f)
    ws = wb.active
    ws["A2"].fill = PatternFill(start_color=JAUNE, end_color=JAUNE, fill_type="solid")
    wb.save(f)

    r = lire_fichier_profs(f)
    assert r.par_code("sortant")[0].nom == "X"


def test_un_classeur_sans_colonne_nom_est_refuse(tmp_path):
    from backend.services.import_profs import lire_fichier_profs

    wb = openpyxl.Workbook()
    wb.active.append(["Truc", "Machin"])
    chemin = tmp_path / "bancal.xlsx"
    wb.save(chemin)

    with pytest.raises(ValueError, match="NOM"):
        lire_fichier_profs(chemin)


def test_fichier_absent(tmp_path):
    from backend.services.import_profs import lire_fichier_profs

    with pytest.raises(ValueError, match="introuvable"):
        lire_fichier_profs(tmp_path / "nexiste-pas.xlsx")


def test_le_tableau_lu_est_conserve(session, tmp_path, annee_factory):
    """Un import charge des données ; il ne les emprunte pas le temps d'un écran."""
    from backend.services.import_profs import (
        enregistrer,
        lire_enregistres,
        lire_fichier_profs,
    )

    annee = annee_factory("2026-2027")
    f = _classeur(tmp_path, [
        ("M.", "ALDRIN", "Thierry", "Technologie", JAUNE),
        ("Mme", "CAREME", "Mégane", "Maths", BLEU),
    ], LEGENDE)

    n = enregistrer(session, lire_fichier_profs(f), annee_id=annee.id)
    assert n == 2

    relus = lire_enregistres(session, annee_id=annee.id)
    assert [(p.nom, p.code) for p in relus] == [
        ("ALDRIN", "sortant"), ("CAREME", "arrivant"),
    ]


def test_un_reimport_remplace_lannee_concernee(session, tmp_path, annee_factory):
    """Un enseignant retiré du classeur ne doit pas y survivre indéfiniment."""
    from backend.services.import_profs import (
        enregistrer,
        lire_enregistres,
        lire_fichier_profs,
    )

    annee = annee_factory("2026-2027")
    enregistrer(session, lire_fichier_profs(_classeur(tmp_path, [
        ("M.", "PARTI", "Jean", "Maths", None),
        ("Mme", "RESTE", "Anne", "SVT", None),
    ], LEGENDE, nom="v1.xlsx")), annee_id=annee.id)

    enregistrer(session, lire_fichier_profs(_classeur(tmp_path, [
        ("Mme", "RESTE", "Anne", "SVT", None),
    ], LEGENDE, nom="v2.xlsx")), annee_id=annee.id)

    assert [p.nom for p in lire_enregistres(session, annee_id=annee.id)] == ["RESTE"]


def test_deux_annees_coexistent(session, tmp_path, annee_factory):
    """Le tableau d'une rentrée s'ajoute, il ne remplace pas le précédent."""
    from backend.services.import_profs import (
        enregistrer,
        lire_enregistres,
        lire_fichier_profs,
    )

    a1 = annee_factory("2025-2026")
    a2 = annee_factory("2026-2027")
    enregistrer(session, lire_fichier_profs(_classeur(tmp_path, [
        ("M.", "ANCIEN", "Paul", "Maths", None)], LEGENDE, nom="a.xlsx")),
        annee_id=a1.id)
    enregistrer(session, lire_fichier_profs(_classeur(tmp_path, [
        ("Mme", "NOUVELLE", "Zoé", "SVT", None)], LEGENDE, nom="b.xlsx")),
        annee_id=a2.id)

    assert [p.nom for p in lire_enregistres(session, annee_id=a1.id)] == ["ANCIEN"]
    assert [p.nom for p in lire_enregistres(session, annee_id=a2.id)] == ["NOUVELLE"]


def test_la_date_dimport_est_connue(session, tmp_path, annee_factory):
    from backend.services.import_profs import (
        date_import,
        enregistrer,
        lire_fichier_profs,
    )

    annee = annee_factory("2026-2027")
    assert date_import(session, annee_id=annee.id) is None

    enregistrer(session, lire_fichier_profs(_classeur(tmp_path, [
        ("M.", "X", "Un", "Maths", None)], LEGENDE)), annee_id=annee.id)
    assert date_import(session, annee_id=annee.id) is not None
