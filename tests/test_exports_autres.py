"""Tests des exports JPM/SmartAir et CardStudio (Lot 11).

L'export PMB n'est plus fabriqué ici : PMB veut treize colonnes dont
sept sont absentes du référentiel. Le programme se contente de répartir
le fichier de Charlemagne — voir `test_repartition_pmb.py`.
"""
from __future__ import annotations

import csv
import io

import openpyxl
import pytest


@pytest.fixture()
def snap_factory(session):
    from backend.models import Snapshot

    def _creer(personne_id, annee_id, **kwargs):
        defaults = {"nom": "MARTIN", "prenom": "Jean", "classe": "3B"}
        defaults.update(kwargs)
        s = Snapshot(personne_id=personne_id, annee_scolaire_id=annee_id, **defaults)
        session.add(s)
        session.commit()
        return s

    return _creer


# ---------------------------------------------------------------------------
# JPM / SmartAir
# ---------------------------------------------------------------------------


def test_jpm_differentiel_a_b_m(session, site_factory, annee_factory, personne_factory, snap_factory):
    from backend.services.exports_jpm import COLONNES_JPM, generer_csv_jpm

    site = site_factory("NDK")
    an_prec = annee_factory("2024-2025")
    an_cour = annee_factory("2025-2026")

    # Ajout (dans cible seulement)
    p_a = personne_factory(site_id=site.id, nom="NEUF", login="neuf")
    snap_factory(p_a.id, an_cour.id, classe="3B")
    # Suppression (dans source seulement)
    p_b = personne_factory(site_id=site.id, nom="SORT", login="sort")
    snap_factory(p_b.id, an_prec.id, classe="TALE")
    # Modification (classe change)
    p_m = personne_factory(site_id=site.id, nom="MOD", login="mod")
    snap_factory(p_m.id, an_prec.id, classe="3B")
    snap_factory(p_m.id, an_cour.id, classe="2NDE")
    # Identique (ignoré)
    p_i = personne_factory(site_id=site.id, nom="IDENT", login="ident")
    snap_factory(p_i.id, an_prec.id, classe="4A")
    snap_factory(p_i.id, an_cour.id, classe="4A")

    contenu, rapport = generer_csv_jpm(
        session=session, site_id=site.id,
        annee_cible_id=an_cour.id, annee_source_id=an_prec.id,
    )

    assert rapport.nb_ajouts == 1
    assert rapport.nb_suppressions == 1
    assert rapport.nb_modifications == 1
    assert rapport.nb_total == 3

    rows = list(csv.DictReader(io.StringIO(contenu.decode("utf-8"))))
    ops = {r["Op"] for r in rows}
    assert ops == {"a", "b", "m"}
    # Colonnes techniques constantes présentes
    for r in rows:
        assert r["Technology"] == "P"
        assert r["Grants"] == "FFFFFF"


def test_jpm_ignore_les_adultes(session, site_factory, annee_factory, personne_factory, snap_factory):
    from backend.services.exports_jpm import generer_csv_jpm

    site = site_factory("NDK")
    an_prec = annee_factory("2024-2025")
    an_cour = annee_factory("2025-2026")

    # Un adulte nouveau — ne doit PAS apparaître dans JPM
    p_adulte = personne_factory(type="adulte", site_id=site.id, login="prof1")
    snap_factory(p_adulte.id, an_cour.id, poste_occupe="ENSEIGNEMENT")

    _, rapport = generer_csv_jpm(
        session=session, site_id=site.id,
        annee_cible_id=an_cour.id, annee_source_id=an_prec.id,
    )
    assert rapport.nb_total == 0


# ---------------------------------------------------------------------------
# CardStudio
# ---------------------------------------------------------------------------


def test_cardstudio_xlsx_avec_13_colonnes(
    session, site_factory, annee_factory, personne_factory, snap_factory
):
    from backend.services.exports_cardstudio import COLONNES_CARDSTUDIO, generer_xlsx_cardstudio

    site = site_factory("NDK")
    annee = annee_factory()
    p = personne_factory(site_id=site.id, nom="DUPONT", prenom="Jean", login="jd")
    snap_factory(p.id, annee.id, classe="3B", niveau="3")

    contenu, rapport = generer_xlsx_cardstudio(
        session=session, site_id=site.id, categorie="tous", annee_cible_id=annee.id,
    )

    wb = openpyxl.load_workbook(io.BytesIO(contenu))
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert header == COLONNES_CARDSTUDIO
    assert rapport.nb_lignes == 1


def test_cardstudio_nom_fichier_photo_par_defaut(
    session, site_factory, annee_factory, personne_factory, snap_factory
):
    """Si le snapshot n'a pas de chemin_photo, on met « NOM Prénom.jpg »."""
    from backend.services.exports_cardstudio import COLONNES_CARDSTUDIO, generer_xlsx_cardstudio

    site = site_factory("NDK")
    annee = annee_factory()
    p = personne_factory(site_id=site.id, nom="DUPONT", prenom="Jean", login="jd")
    snap_factory(p.id, annee.id, classe="3B", chemin_photo=None)

    contenu, _ = generer_xlsx_cardstudio(
        session=session, site_id=site.id, categorie="tous", annee_cible_id=annee.id,
    )
    wb = openpyxl.load_workbook(io.BytesIO(contenu))
    row = list(wb.active.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    idx = COLONNES_CARDSTUDIO.index("NomFichierPhoto")
    assert row[idx] == "DUPONT Jean.jpg"


def test_cardstudio_categorie_nouveaux(
    session, site_factory, annee_factory, personne_factory, snap_factory
):
    from backend.services.exports_cardstudio import generer_xlsx_cardstudio

    site = site_factory("NDK")
    an_prec = annee_factory("2024-2025")
    an_cour = annee_factory("2025-2026")
    p_reste = personne_factory(site_id=site.id, login="reste")
    snap_factory(p_reste.id, an_prec.id)
    snap_factory(p_reste.id, an_cour.id)
    p_neuf = personne_factory(site_id=site.id, login="neuf")
    snap_factory(p_neuf.id, an_cour.id)

    _, r = generer_xlsx_cardstudio(
        session=session, site_id=site.id, categorie="nouveaux",
        annee_cible_id=an_cour.id, annee_source_id=an_prec.id,
    )
    assert r.nb_lignes == 1


def test_cardstudio_ignore_les_adultes(session, site_factory, annee_factory, personne_factory, snap_factory):
    from backend.services.exports_cardstudio import generer_xlsx_cardstudio

    site = site_factory("NDK")
    annee = annee_factory()
    p_adulte = personne_factory(type="adulte", site_id=site.id, login="prof1")
    snap_factory(p_adulte.id, annee.id)

    _, r = generer_xlsx_cardstudio(
        session=session, site_id=site.id, categorie="tous", annee_cible_id=annee.id,
    )
    assert r.nb_lignes == 0
